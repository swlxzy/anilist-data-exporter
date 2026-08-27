import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Install it with:")
    print("pip install openpyxl")
    sys.exit(1)


def safe_filename(name):
    invalid = '<>:"/\\|?*'
    for char in invalid:
        name = name.replace(char, "_")
    return name.strip() or "export"


def load_json(path):
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def collect_columns(data):
    columns = ["title", "score", "average"]

    optional_columns = [
        "started",
        "completed",
        "notes",
        "advanced",
        "emptyAdvanced"
    ]

    for lists in data.get("lists", {}).values():
        for item in lists:
            for column in optional_columns:
                if column in item and column not in columns:
                    columns.append(column)

    return columns


def format_value(value):
    if value is None:
        return ""

    if isinstance(value, dict):
        return "\n".join(
            f"{key}: {value[key]}"
            for key in value
        )

    if isinstance(value, list):
        return ", ".join(str(item) for item in value)

    return value


def create_workbook(data):
    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    columns = collect_columns(data)

    for list_name, entries in data.get("lists", {}).items():
        sheet_name = safe_filename(list_name)[:31] or "List"

        original_name = sheet_name
        counter = 2

        while sheet_name in workbook.sheetnames:
            suffix = f"_{counter}"
            sheet_name = original_name[:31 - len(suffix)] + suffix
            counter += 1

        worksheet = workbook.create_sheet(sheet_name)

        for column_index, column in enumerate(columns, start=1):
            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=column
            )
            cell.font = Font(bold=True)

        for row_index, item in enumerate(entries, start=2):
            for column_index, column in enumerate(columns, start=1):
                value = format_value(item.get(column))

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )

                if isinstance(value, str) and "\n" in value:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

        for column_cells in worksheet.columns:
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    length = max(
                        len(line)
                        for line in str(cell.value).split("\n")
                    )
                    max_length = max(max_length, length)

            width = min(max_length + 2, 50)
            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = width

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    return workbook


def main():
    print("AniList JSON to Excel Converter")
    print("--------------------------------")

    json_path = input("\nJSON file path: ").strip()

    if not json_path:
        print("Error: No file path provided.")
        return

    if not os.path.isfile(json_path):
        print("Error: File not found.")
        return

    if not json_path.lower().endswith(".json"):
        print("Error: The selected file is not a JSON file.")
        return

    try:
        data = load_json(json_path)
    except json.JSONDecodeError:
        print("Error: Invalid JSON file.")
        return
    except OSError as error:
        print(f"Error reading file: {error}")
        return

    if not isinstance(data, dict) or "lists" not in data:
        print("Error: This does not appear to be an AniList Data Exporter file.")
        return

    workbook = create_workbook(data)

    base_name = os.path.splitext(os.path.basename(json_path))[0]
    output_path = os.path.join(
        os.path.dirname(json_path),
        safe_filename(base_name) + ".xlsx"
    )

    counter = 2

    while os.path.exists(output_path):
        output_path = os.path.join(
            os.path.dirname(json_path),
            f"{safe_filename(base_name)}_{counter}.xlsx"
        )
        counter += 1

    try:
        workbook.save(output_path)
    except OSError as error:
        print(f"Error saving Excel file: {error}")
        return

    print("\nConversion completed!")
    print(f"Output file: {os.path.basename(output_path)}")
    print(f"Saved to: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()